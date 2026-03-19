import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io
import re

# ==============================
# CONFIGURAÇÃO DA PÁGINA
# ==============================
st.set_page_config(
    page_title="Auditoria FATURA x APEX",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
.stApp { background-color: #f4f6fb; }
h1, h2, h3 { color: #ff8c00; }
div[data-testid="stMetric"] {
    background-color: white;
    padding: 20px;
    border-radius: 12px;
    border-left: 6px solid #ff8c00;
    box-shadow: 0px 4px 12px rgba(0,0,0,0.08);
}
.stButton>button {
    background-color: #ff8c00;
    color: white;
    border-radius: 8px;
    border: none;
}
</style>
""", unsafe_allow_html=True)

st.title("📊 Auditoria FATURA x APEX")
st.caption("Carregue os arquivos abaixo para iniciar a comparação.")
st.divider()


# ==============================
# FUNÇÕES AUXILIARES
# ==============================

def converter_valor(valor):
    """Replica ConverterValorParaNumero do VBA."""
    try:
        s = str(valor)
        s = s.replace("R$", "").replace(" ", "").replace(".", "")
        s = s.replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def limpar_codigo(valor):
    """Remove .0 de inteiros lidos como float pelo pandas."""
    s = str(valor).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def extrair_codigo_fatura(pat, serie):
    """
    Replica lógica do VBA:
    1º usa PAT (col A); se vazio, usa Num Série (col C) só se for número puro.
    """
    pat   = limpar_codigo(pat)   if pd.notna(pat)   else ""
    serie = limpar_codigo(serie) if pd.notna(serie) else ""

    if pat and pat not in ("", "nan", "0"):
        try:
            return str(int(float(pat)))
        except Exception:
            return pat

    if serie and serie not in ("", "nan", "0"):
        if re.fullmatch(r"\d+", serie):
            try:
                return str(int(float(serie)))
            except Exception:
                return serie

    return ""


def comparar_fatura_apex(df_fatura, df_apex):
    """Replica exatamente a macro VBA CompararFaturaApex."""

    # Dicionário FATURA  (col A=PAT, col B=Tot Geral, col C=Num Série)
    dict_fatura = {}
    for _, row in df_fatura.iterrows():
        pat   = row.iloc[0]
        valor = row.iloc[1]
        serie = row.iloc[2] if len(row) > 2 else ""
        codigo = extrair_codigo_fatura(pat, serie)
        if codigo:
            dict_fatura[codigo] = converter_valor(valor)

    # Dicionário APEX  (col A=Tombo, col B=Vr Loc)
    dict_apex = {}
    for _, row in df_apex.iterrows():
        codigo = limpar_codigo(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        if codigo and codigo not in ("", "nan"):
            dict_apex[codigo] = converter_valor(row.iloc[1])

    # Comparação — mesma ordem do VBA
    resultado = []
    dict_apex_restante = dict(dict_apex)

    for codigo, valor_f in dict_fatura.items():
        if codigo in dict_apex_restante:
            valor_a = dict_apex_restante.pop(codigo)
            status  = "Valores iguais" if abs(valor_f - valor_a) < 0.01 else "Valores diferentes"
        else:
            valor_a = None
            status  = "Só na Fatura Verificar APEX" if valor_f > 0 else "Só na FATURA"

        resultado.append({
            "Código":       codigo,
            "Valor FATURA": valor_f,
            "Valor APEX":   valor_a,
            "Status":       status
        })

    # Sobras do APEX
    for codigo, valor_a in dict_apex_restante.items():
        resultado.append({
            "Código":       codigo,
            "Valor FATURA": None,
            "Valor APEX":   valor_a,
            "Status":       "Só na APEX"
        })

    return pd.DataFrame(resultado)


def gerar_excel(df_resultado, df_fatura_raw, df_apex_raw):
    """Excel com abas Comparacao, FATURA e APEX — cores idênticas ao VBA."""
    wb = Workbook()

    cores_hex = {
        "Valores iguais":              "C6EFCE",
        "Valores diferentes":          "FFC7CE",
        "Só na FATURA":                "FFEB9C",
        "Só na APEX":                  "FFEB9C",
        "Só na Fatura Verificar APEX": "FFC000",
    }

    header_fill = PatternFill("solid", fgColor="FF8C00")
    header_font = Font(bold=True, color="FFFFFF")

    # Aba Comparacao
    ws_res = wb.active
    ws_res.title = "Comparacao"
    colunas = ["Código", "Valor FATURA", "Valor APEX", "Status"]
    for c, h in enumerate(colunas, 1):
        cell = ws_res.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(df_resultado.itertuples(index=False), 2):
        status = row.Status
        fill   = PatternFill("solid", fgColor=cores_hex.get(status, "FFFFFF"))
        for c, value in enumerate(row, 1):
            cell = ws_res.cell(row=r, column=c, value=value)
            if c == 4:
                cell.fill = fill

    for col in ws_res.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_res.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Aba FATURA
    ws_fat = wb.create_sheet("FATURA")
    for c, h in enumerate(["PAT", "Tot Geral", "Num Série"], 1):
        ws_fat.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, row in enumerate(df_fatura_raw.itertuples(index=False), 2):
        for c, v in enumerate(row, 1):
            ws_fat.cell(row=r, column=c, value=v)

    # Aba APEX
    ws_apex = wb.create_sheet("APEX")
    for c, h in enumerate(["Tombo", "Vr Loc"], 1):
        ws_apex.cell(row=1, column=c, value=h).font = Font(bold=True)
    for r, row in enumerate(df_apex_raw.itertuples(index=False), 2):
        for c, v in enumerate(row, 1):
            ws_apex.cell(row=r, column=c, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==============================
# UPLOADS
# ==============================
col1, col2 = st.columns(2)

with col1:
    st.markdown("#### 📄 Arquivo FATURA")
    st.caption("Colunas usadas: **A** = PAT · **B** = Tot Geral · **C** = Num Série")
    file_fatura = st.file_uploader("Selecione o arquivo da FATURA",
                                   type=["xlsx","xlsm","xls"], key="fatura")

with col2:
    st.markdown("#### 📄 Arquivo APEX")
    st.caption("Colunas usadas: **A** = Tombo · **B** = Vr Loc")
    file_apex = st.file_uploader("Selecione o arquivo do APEX",
                                 type=["xlsx","xlsm"], key="apex")

st.divider()

# ==============================
# PROCESSAMENTO
# ==============================
if file_fatura and file_apex:

    with st.spinner("Comparando dados..."):

        df_fat_full  = pd.read_excel(file_fatura, header=0)
        df_fatura_dest = df_fat_full.iloc[:, [0, 1, 2]].copy()
        df_fatura_dest.columns = ["PAT", "Tot Geral", "Num Série"]

        df_apex_full = pd.read_excel(file_apex, header=0)
        df_apex_dest = df_apex_full.iloc[:, [0, 1]].copy()
        df_apex_dest.columns = ["Tombo", "Vr Loc"]

        df_resultado = comparar_fatura_apex(df_fatura_dest, df_apex_dest)

    # MÉTRICAS
    total      = len(df_resultado)
    iguais     = (df_resultado["Status"] == "Valores iguais").sum()
    diferentes = (df_resultado["Status"] == "Valores diferentes").sum()
    so_fatura  = df_resultado["Status"].isin(["Só na FATURA","Só na Fatura Verificar APEX"]).sum()
    so_apex    = (df_resultado["Status"] == "Só na APEX").sum()

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total",                  total)
    m2.metric("✅ Valores Iguais",       iguais)
    m3.metric("❌ Valores Diferentes",   diferentes)
    m4.metric("⚠️ Só na FATURA",         so_fatura)
    m5.metric("🔵 Só na APEX",           so_apex)

    st.divider()

    # GRÁFICOS
    cores_graf = {
        "Valores iguais":              "#2ecc71",
        "Valores diferentes":          "#e74c3c",
        "Só na FATURA":                "#f39c12",
        "Só na APEX":                  "#3498db",
        "Só na Fatura Verificar APEX": "#e67e22",
    }

    colA, colB = st.columns(2)
    with colA:
        graf = df_resultado["Status"].value_counts().reset_index()
        graf.columns = ["Status","Quantidade"]
        fig = px.bar(graf, x="Status", y="Quantidade", color="Status",
                     text="Quantidade", color_discrete_map=cores_graf,
                     title="Distribuição de Status")
        fig.update_layout(showlegend=False)
        st.plotly_chart(fig, use_container_width=True)

    with colB:
        fig2 = px.scatter(
            df_resultado.dropna(subset=["Valor FATURA","Valor APEX"]),
            x="Valor FATURA", y="Valor APEX", color="Status",
            color_discrete_map=cores_graf, hover_data=["Código"],
            title="Comparação FATURA x APEX"
        )
        st.plotly_chart(fig2, use_container_width=True)

    st.divider()

    # FILTROS
    colF1, colF2 = st.columns([2, 2])
    busca      = colF1.text_input("🔎 Buscar por Código")
    status_sel = colF2.radio("Filtrar Status",
                             ["Todos"] + sorted(df_resultado["Status"].unique().tolist()),
                             horizontal=True)

    df_view = df_resultado.copy()
    if busca:
        df_view = df_view[df_view["Código"].astype(str).str.contains(busca, case=False)]
    if status_sel != "Todos":
        df_view = df_view[df_view["Status"] == status_sel]

    # TABELA
    cores_tabela = {
        "Valores iguais":              "#eafaf1",
        "Valores diferentes":          "#fdecea",
        "Só na FATURA":                "#fef9e7",
        "Só na APEX":                  "#eaf4fb",
        "Só na Fatura Verificar APEX": "#fff3cd",
    }

    def color_status(row):
        cor = cores_tabela.get(row["Status"], "")
        return [f"background-color:{cor}"] * len(row)

    st.subheader("📋 Resultado da Comparação")
    st.dataframe(df_view.style.apply(color_status, axis=1),
                 use_container_width=True, height=500)

    st.divider()

    # DOWNLOADS
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button("📥 Baixar resultado (CSV)",
                           df_view.to_csv(index=False).encode("utf-8"),
                           file_name="auditoria_resultado.csv", mime="text/csv")
    with dl2:
        excel_buf = gerar_excel(df_resultado, df_fatura_dest, df_apex_dest)
        st.download_button("📥 Baixar Excel completo (3 abas)", excel_buf,
                           file_name="auditoria_completa.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("⬆️ Carregue os dois arquivos acima para iniciar a auditoria.")
